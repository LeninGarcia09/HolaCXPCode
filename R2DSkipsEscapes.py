# Import required libraries for data processing, Excel manipulation, and GUI file selection
import pandas as pd  # For CSV data manipulation and analysis
import openpyxl  # For Excel file creation and formatting
from openpyxl.styles import Font, PatternFill, Border, Side  # For Excel cell styling
from tkinter import Tk, filedialog  # For creating file selection dialog


def select_csv_file():
    """
    Opens a file dialog for the user to select a CSV file.

    Returns:
        str: Full path to the selected CSV file, or empty string if cancelled
    """
    # Create a Tkinter root window (required for file dialogs)
    root = Tk()
    # Hide the main window - we only want the file dialog to appear
    root.withdraw()
    # Open file selection dialog filtered to show only CSV files
    file_path = filedialog.askopenfilename(
        filetypes=[("CSV Files", "*.csv")],
        title="Select CSV File"
    )
    return file_path


def import_csv_and_prepare_data(wb_path):
    """
    Prompts user to select a CSV file, imports it as a DataFrame,
    and adds it as a new sheet to the Excel workbook.

    Args:
        wb_path (str): Path to the Excel workbook file

    Returns:
        pandas.DataFrame: The imported CSV data, or None if no file selected
    """
    # Get CSV file path from user selection
    file_path = select_csv_file()
    if not file_path:
        return None

    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(file_path)

    # Add the DataFrame as a new sheet in the existing Excel workbook
    with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='R2D IcM Data', index=False)

    return df


def create_grouped_worksheet(wb_path, df, col_name, sheet_name, group_values=None):
    """
    Creates a new worksheet with data grouped by a specific column.
    Each group is separated by a header row and blank row for readability.

    Args:
        wb_path (str): Path to the Excel workbook file
        df (pandas.DataFrame): The source data to group
        col_name (str): Name of the column to group by
        sheet_name (str): Name for the new worksheet
        group_values (list, optional): Specific values to group by. If None, uses all unique values.
    """
    # Determine which values to group by
    if group_values is None or len(group_values) == 0:
        # Get all unique values from the column, replacing NaN with '(Blank)'
        group_values = df[col_name].fillna('(Blank)').unique()
    else:
        # Replace empty strings with '(Blank)' for better readability
        group_values = [(v if v != "" else "(Blank)") for v in group_values]

    # Load the existing workbook and create a new worksheet
    wb = openpyxl.load_workbook(wb_path)
    ws = wb.create_sheet(sheet_name)

    # Add the header row with column names
    header = list(df.columns)
    ws.append(header)

    # Process each group separately
    for group in group_values:
        # Add a group header row (e.g., "R2DFranchiseName = GroupName")
        ws.append([f"{col_name} = {group}"] + [""] * (len(header) - 1))

        # Add all rows that belong to this group
        for row in df[df[col_name].fillna('(Blank)') == group].itertuples(index=False):
            ws.append(list(row))

        # Add a blank row to separate groups visually
        ws.append([""] * len(header))

    # Save the workbook with the new worksheet
    wb.save(wb_path)


def generate_outage_reports(wb_path, df):
    """
    Creates a formatted worksheet with data sorted by severity column.
    Applies professional styling including header formatting and alternating row colors.

    Args:
        wb_path (str): Path to the Excel workbook file
        df (pandas.DataFrame): The source data to sort and format
    """
    sheet_name = "Organized by Severity"

    # Sort the DataFrame by the 4th column (index 3), assumed to be Severity
    df_sorted = df.sort_values(by=df.columns[3])  # Assuming 4th column is Severity

    # Load workbook and create new worksheet
    wb = openpyxl.load_workbook(wb_path)
    ws = wb.create_sheet(sheet_name)

    # Add header row and all sorted data rows
    ws.append(list(df.columns))
    for row in df_sorted.itertuples(index=False):
        ws.append(list(row))

    # Apply professional formatting to the header row
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")  # Blue background
    header_font = Font(bold=True, color="FFFFFF")  # White, bold text
    for cell in ws[1]:  # First row (header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(bottom=Side(style='thin'))  # Add bottom border

    # Apply alternating row colors for better readability
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Apply light gray fill to even-numbered rows
        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill

    # Save the formatted workbook
    wb.save(wb_path)


def process_csv_and_format():
    """
    Main function that orchestrates the entire CSV processing workflow.

    Process:
    1. Creates or opens the output Excel workbook
    2. Prompts user to select and import CSV data
    3. Generates multiple organized worksheets with different groupings
    4. Applies formatting to make reports professional and readable
    """
    wb_path = "output.xlsx"

    # Create a new workbook if it doesn't exist, or load existing one
    try:
        wb = openpyxl.load_workbook(wb_path)
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        wb = openpyxl.Workbook()
        wb.save(wb_path)

    # Import CSV data and add it to the workbook
    df = import_csv_and_prepare_data(wb_path)
    if df is None:
        print("No file selected.")
        return

    # Generate the main formatted report sorted by severity
    generate_outage_reports(wb_path, df)

    # Create specialized worksheets grouped by different categories

    # Group by franchise name (using all unique values in the column)
    create_grouped_worksheet(wb_path, df, "R2D Franchise Name", "Franchise_Categories")

    # Group by lease request status (TRUE/FALSE/empty values)
    create_grouped_worksheet(wb_path, df, "Is Lease Request?", "Lease_Categories", ["TRUE", "FALSE", ""])

    # Group by manual touch requirement (TRUE/FALSE/empty values)
    create_grouped_worksheet(wb_path, df, "Manual Touch", "Manual_Touch_Categories", ["Yes", "No", ""])

    # Group by R2D skip/escape status (specific predefined categories)
    create_grouped_worksheet(wb_path, df, "R2D Skip Escape", "R2D_Skip_Escape_Categories",
                             ["Escape", "Skip", ""])

    #Group by R2D R2D Assessed Risk (High, Medium, Low)
    create_grouped_worksheet(wb_path, df, "R2D Assessed Risk", "R2D_Assessed_Risk_Categories",
                             ["High", "Medium", "Low", ""])

    print("El trabajo de procesamiento se ha completado. Check output.xlsx.")


if __name__ == "__main__":
    """
    Entry point of the script. 
    This ensures the main function only runs when the script is executed directly,
    not when imported as a module.
    """
    process_csv_and_format()